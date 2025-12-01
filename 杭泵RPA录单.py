import json
import re
import os
import copy
from openpyxl import load_workbook
from datetime import datetime, timedelta

def excel_date_to_str(value):
    """将Excel日期或字符串日期统一转换为 2025.11.07 格式"""
    if not value:
        return ""

    # 数字日期（Excel序号）
    if isinstance(value, (int, float)):
        date = datetime(1899, 12, 30) + timedelta(days=value)
        return date.strftime("%Y.%m.%d")

    # datetime 类型
    if isinstance(value, datetime):
        return value.strftime("%Y.%m.%d")

    # 字符串类型，尝试匹配中文或横线日期
    if isinstance(value, str):
        text = value.strip()
        # 中文年月日
        m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
        if m:
            y, mth, d = m.groups()
            return f"{int(y):04d}.{int(mth):02d}.{int(d):02d}"
        # 横线或斜杠格式 2025-11-10 / 2025/11/10
        m = re.match(r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})", text)
        if m:
            y, mth, d = m.groups()
            return f"{int(y):04d}.{int(mth):02d}.{int(d):02d}"
        # 否则直接返回
        return text

    return str(value)

def find_info_row(ws, start_row):
    """
    找到编号/客户名称所在的行（物料下面第一行非物料号）
    """
    row = start_row + 1
    while row <= ws.max_row:
        c = str(ws[f"C{row}"].value or "").strip()
        a = str(ws[f"A{row}"].value or "").strip()

        # 物料号区：C列有值；信息区：C列为空而A列有文本
        if c == "" and a != "":
            return row
        row += 1

    return start_row + 2  #保底，兼容你当前模板



def get_merged_value(ws, cell_ref):
    """读取合并单元格值"""
    cell = ws[cell_ref]
    if cell.value is not None:
        return str(cell.value).replace("\n", "")
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            val = ws.cell(merged.min_row, merged.min_col).value
            return str(val).replace("\n", "").replace(" ","") if val else ""
    return ""

def find_all_block_starts(ws):
    """自动扫描所有子订单起始行"""
    block_starts = []
    for row in range(1, ws.max_row + 1):
        val = ws[f"C{row}"].value
        if isinstance(val, str) and re.search(r"物料 *号|物料编 *号", val):
            block_starts.append(row)
    return block_starts

def parse_order_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    all_orders = []
    block_starts = find_all_block_starts(ws)

    for start_row in block_starts:
        info_row = find_info_row(ws, start_row)
        bianhao = get_merged_value(ws, f"A{info_row}")
        name = get_merged_value(ws, f"B{info_row}")
        receiver = get_merged_value(ws, f"J{start_row + 1}")


        order_data = {
            "订单类型": "",
            "销售组织": "3900",
            "分销渠道": "10",
            "产品组": "10",
            "销售组": "330",
            "编号": bianhao,
            "名称": name,
            "收货人信息": receiver,
            "最久交期":"",
            "是否安装调试": "",
            "备注1": "",
            "备注2": "",
            "items": []
        }

        # 表格中数据从表头下一行开始
        row = start_row + 1
        while row <= ws.max_row:
            text = str(ws[f"C{row}"].value or "").strip()

            # 遇到 "是否安装调试验收" 说明子订单结束
            if "是否安装调试验收" in text:
                next_val = get_merged_value(ws, f"C{row + 1}")
                order_data["是否安装调试"] = next_val
                # 开始往下找 D 列备注1 / 备注2
                for i in range(row, row + 3):  # 往下查几行范围
                    d_val = str(ws[f"D{i}"].value or "").strip()
                    d_val = d_val.replace(":", "：")
                    if "备注1" in d_val:
                        # split("：", 1) 表示只切分第一个冒号，[1] 取冒号后面的所有内容
                        parts = d_val.split("：", 1)
                        order_data["备注1"] = parts[1].strip() if len(parts) > 1 else ""
                    if "备注2" in d_val:
                        parts = d_val.split("：", 1)
                        order_data["备注2"] = parts[1].strip() if len(parts) > 1 else ""
                break

            # 跳过空行或非物料号
            if not text:
                row += 1
                continue

            item = {
                "物料号": get_merged_value(ws, f"C{row}"),
                "水泵型号": get_merged_value(ws, f"D{row}"),
                "工厂": get_merged_value(ws, f"E{row}"),
                "数量": get_merged_value(ws, f"F{row}"),
                "单价": get_merged_value(ws, f"G{row}"),
                "金额": get_merged_value(ws, f"H{row}"),
            }
            # 获取交期列，并格式化日期
            raw_jiaoqi = ws[f"I{row}"].value
            item["交期"] = excel_date_to_str(raw_jiaoqi)

            order_data["items"].append(item)
            row += 1

        all_orders.append(order_data)

    # 第二阶段：仅当子订单中存在工厂为3900的物料时进行拆分
    grouped_orders = []
    for order in all_orders:
        items = order["items"]
        has_3900 = any(item["工厂"] == "3900" for item in items)

        if has_3900:
            # 拆成两个子订单：一个是3900，一个是其他工厂
            items_3900 = [i for i in items if i["工厂"] == "3900"]
            items_other = [i for i in items if i["工厂"] != "3900"]

            # 工厂=3900的单独一个子订单
            if items_3900:
                new_order_3900 = copy.deepcopy(order)
                new_order_3900["items"] = items_3900
                grouped_orders.append(new_order_3900)

            # 其他工厂的全部合并为一个
            if items_other:
                new_order_other = copy.deepcopy(order)
                new_order_other["items"] = items_other
                grouped_orders.append(new_order_other)
        else:
            # 没有3900工厂，不拆分
            grouped_orders.append(order)

    # 计算最久交期
    for order in grouped_orders:
        latest = ""
        for item in order["items"]:
            val = item.get("交期", "")
            if val and val.lower() != "none":
                if latest == "" or val > latest:
                    latest = val
        order["最久交期"] = latest


    # 第三阶段：根据工厂编号判断订单类型
    for order in grouped_orders:
        factories = {str(item.get("工厂", "")).strip() for item in order["items"]}
        factories = {f for f in factories if f}  # 去掉空白
        order["订单类型"] = ""  # 默认空

        # 没有工厂信息，直接标记为未知
        if not factories:
            print(f"工厂为空，无法判断订单类型: 编号={order['编号']}")
            order["订单类型"] = ""
            continue

        if factories == {"3900"}:
            order["订单类型"] = "Z001"
        elif factories.issubset({"1073", "1079", "3520"}):
            order["订单类型"] = "Z007"
        else:
            print(f"未知的工厂编号: {factories}")

    return grouped_orders


if __name__ == "__main__":
    file_path = r"D:\青臣云起\项目\南方流体模板解析\文件\杭泵成都办2025-12-01-04四川比佰特环保科技有限公司+发货单  .xlsx"
    json数据解析 = parse_order_excel(file_path)
    json_path = os.path.join(os.path.dirname(file_path), "json数据解析.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json数据解析, f, ensure_ascii=False, indent=4)
    print("解析完成")

