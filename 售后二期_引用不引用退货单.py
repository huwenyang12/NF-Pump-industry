import json
import re
import os
from openpyxl import load_workbook


def get_merged_value(ws, cell_ref):
    """读取合并单元格值（取合并区域左上角）"""
    cell = ws[cell_ref]
    if cell.value is not None:
        return str(cell.value).replace("\n", "").replace("\r", "").strip()
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            val = ws.cell(merged.min_row, merged.min_col).value
            return str(val).replace("\n", "").replace("\r", "").strip() if val is not None else ""
    return ""


def parse_serial_list(raw: str):
    """把序列号单元格内容解析成列表（逗号/分号/换行/空格均可分隔，并去重保序）"""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    s = s.replace("，", ",").replace("；", ",").replace(";", ",")
    s = s.replace("\n", ",").replace("\r", ",").replace(" ", ",")
    parts = [x.strip() for x in s.split(",") if x.strip()]
    seen = set()
    uniq = []
    for x in parts:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq



def parse_global_footer(ws):
    footer = {"是否安装调试验收": "", "备注1": "", "备注2": ""}
    # 1) 先扫 G 列找 “是否安装调试验收”
    for r in range(1, ws.max_row + 1):
        g = get_merged_value(ws, f"G{r}")
        if g and "是否安装调试验收" in g:
            footer["是否安装调试验收"] = get_merged_value(ws, f"G{r+1}")
            break
    # 2) 再扫 H 列找 “备注1/备注2”
    for r in range(1, ws.max_row + 1):
        h = get_merged_value(ws, f"H{r}")
        if h:
            if (not footer["备注1"]) and ("备注1" in h):
                footer["备注1"] = get_merged_value(ws, f"I{r}")
            if (not footer["备注2"]) and ("备注2" in h):
                footer["备注2"] = get_merged_value(ws, f"I{r}")
        if footer["备注1"] and footer["备注2"]:
            break
    return footer



def parse_order_excel(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    # 全局尾部字段（整单一次）
    footer = parse_global_footer(ws)
    # 用 dict 聚合：key -> order_data
    grouped = {}  # {(order_no, sold_to): order_data}
    for row in range(1, ws.max_row + 1):
        g_val = get_merged_value(ws, f"G{row}")  # 物料号列（核心列）
        # 只处理“物料号是纯数字串”的行
        if not g_val:
            continue
        if re.search(r"物料\s*号|物料\s*编\s*号", g_val):
            continue
        if not re.fullmatch(r"\d{6,30}", g_val):
            continue
        # ====== 订单头字段（同一订单通常是合并单元格，直接取当前行即可）======
        quote = str(get_merged_value(ws, f"A{row}") or "").strip()  # 是否引用
        no = get_merged_value(ws, f"B{row}")                        # 订单号
        sold_to = get_merged_value(ws, f"C{row}")                   # 售达方编码
        office = get_merged_value(ws, f"D{row}")                    # 销售办事处
        customer_ref = get_merged_value(ws, f"E{row}")              # 客户参考/抬头

        group_key = (str(no or "").strip(), str(sold_to or "").strip())

        # 初始化订单对象（第一次见到这个 key）
        if group_key not in grouped:
            grouped[group_key] = {
                "是否引用": quote,
                "订单类型": "Z008",
                "销售组织": "1072",
                "分销渠道": "10",
                "产品组": "10",
                "销售办事处": office,
                "销售组": "270",
                "订单号": no,
                "售达方编码": sold_to,
                "客户参考": customer_ref,

                # 整单一次的尾部字段：每个订单都带上（更稳妥）
                "是否安装调试验收": footer.get("是否安装调试验收", ""),
                "备注1": footer.get("备注1", ""),
                "备注2": footer.get("备注2", ""),

                "退货行号列表": [],
                "items": [],
            }
        else:
            # 同 key 的订单头字段如果前面为空，这里补一次（防止有些行合并导致读取为空）
            od = grouped[group_key]
            if not od.get("是否引用") and quote:
                od["是否引用"] = quote
            if not od.get("销售办事处") and office:
                od["销售办事处"] = office
            if not od.get("客户参考") and customer_ref:
                od["客户参考"] = customer_ref

        # ====== 明细行 ======
        item = {
            "项目行号": get_merged_value(ws, f"F{row}"),
            "物料号": get_merged_value(ws, f"G{row}"),
            "配件型号": get_merged_value(ws, f"H{row}"),
            "序列号": parse_serial_list(get_merged_value(ws, f"I{row}")),
            "订单数量": get_merged_value(ws, f"J{row}"),
            "金额": get_merged_value(ws, f"K{row}"),
            "工厂": get_merged_value(ws, f"L{row}"),
            "采购订单编号": get_merged_value(ws, f"M{row}"),
            "拒绝原因": get_merged_value(ws, f"N{row}"),
            "显示抬头详细信息": get_merged_value(ws, f"O{row}"),
            "物料文本": get_merged_value(ws, f"P{row}"),
        }

        grouped[group_key]["items"].append(item)

        line_no = str(item.get("项目行号", "") or "").strip()
        if line_no:
            grouped[group_key]["退货行号列表"].append(line_no)

    # ====== 输出 list，并做退货行号去重排序 ======
    all_orders = []
    for _, order_data in grouped.items():
        if not order_data["items"]:
            continue

        # 去重并按数字排序退货行号列表
        seen = set()
        uniq = []
        for x in order_data["退货行号列表"]:
            if x not in seen:
                seen.add(x)
                uniq.append(x)
        try:
            uniq.sort(key=lambda s: int(re.sub(r"\D", "", s) or 0))
        except Exception:
            pass
        order_data["退货行号列表"] = uniq

        all_orders.append(order_data)

    return all_orders


if __name__ == "__main__":
    file_path = r"D:\\青臣云起\\项目\\南方流体模板解析\\文件\\20260113引用不引用退货单模板--订单.xlsx"
    data = parse_order_excel(file_path)

    json_path = os.path.join(os.path.dirname(file_path), "json数据解析.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    print("解析文件完成，订单数：", len(data))
