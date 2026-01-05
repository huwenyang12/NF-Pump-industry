import datetime
import pandas as pd
import os
import requests
import datetime
import time
import log,shutil
from openpyxl import Workbook,load_workbook
import os
import re, util
from exceptions import *
import json

def parse_result(deliDateList, itemList):
    '''
    for dd in deliDateList:
        dd=re.sub('[\u4e00-\u9fa5]', '', dd)
        if ('.' in dd or '-' in dd) and not dd.startswith('202'):
            dd=datetime.datetime.now().strftime('%Y') + '.' + dd
        fdt = getDate(dd)
        for il in itemList:
            if il["交期"] == dd:
                il["交期"] = fdt
    '''
    return itemList

#解析录单文件
def parse(file):
    df = pd.read_excel(file, sheet_name=0)

    result_items = []

    deliDateList = []
    itemList = []
    dataStart = False
    lastId = None
    lastCustomer = None
    lastDelDate = None
    lastReceiver = None
    contractNo = None
    ps1=""
    ps2=""
    num=0
    num2=0
    num3=0
    num4=0
    has_seperate_beizhu1 = False
    委托编号 = ""
    for idx, row in df.iterrows():
        try:
            if str(df.loc[idx, "Unnamed: 10"]).__contains__("备注1"):
                has_seperate_beizhu1 = True
        except:
            pass
        料号 = str(df.loc[idx, "Unnamed: 2"]).strip()
        if( 料号 == "物料号"):
            dataStart = True
            if len(deliDateList) > 0:
                itemList = parse_result(deliDateList, itemList)
                #根据成品标准周期 判断是否要替换工厂和交期
                #util.update_date(itemList)
                result_items.append(itemList)
            deliDateList = []
            itemList = []
            lastId = None
            lastCustomer = None
            lastDelDate = None
            lastReceiver = None
            contractNo = None
            ps1=""
            ps2=""
            num=0
            num2=0
            num3=0
            num4=0
            continue
        if 料号=="物料编码" or 料号=="料号":
            raise ExcelDataException("1001","表头不符合物料号")
        if(dataStart == False):
            continue
        raw_serial = str(df.loc[idx, "Unnamed: 4"]).strip()
        if raw_serial.lower() == "nan" or raw_serial == "":
            序列号列表 = []
        else:
            raw_serial = raw_serial.replace("，", ",")  # 防止中文逗号
            序列号列表 = [s.strip() for s in raw_serial.split(",") if s.strip()]
        
        编号 = str(df.loc[idx, row.index[0]])
        备注1= str(df.loc[idx, row.index[3]])
        if str(df.loc[idx, row.index[2]]) == "是否安装调试验收":
            anzhuang_flag = str(df.loc[idx+1, row.index[2]])
            while num3<itemList.__len__():
                itemList[num3]["是否安装调试验收"]=anzhuang_flag
                num3+=1
        else:
            anzhuang_flag = ''
        备注2= str(df.loc[idx, row.index[3]])
        客户组= str(df.loc[idx, row.index[10]])
        if row.size > 11:
            新办事处= str(df.loc[idx, row.index[11]])
        else:
            新办事处 = ''
        客户 = str(df.loc[idx, "Unnamed: 1"])
        工厂=  str(df.loc[idx, "Unnamed: 5"])
        数量 = str(df.loc[idx, "Unnamed: 6"])
        单价 = str(df.loc[idx, "Unnamed: 7"])
        dttt = df.loc[idx, "Unnamed: 9"]
        if pd.isna(dttt):
            交期 = None
        elif isinstance(dttt, (int, float)):
            # Excel日期序号转换
            交期 = (datetime.datetime(1900,1,1) + datetime.timedelta(days=int(dttt)-2)).strftime("%Y.%m.%d")
        else:
            交期 = str(dttt).strip()
            if 交期.lower() == 'nan' or not 交期:
                交期 = None

        if has_seperate_beizhu1:
            备注1 = str(df.loc[idx, "Unnamed: 10"])
            收货人 = str(df.loc[idx, "Unnamed: 11"])
            if str(df.loc[idx, "Unnamed: 12"]) not in ['', 'nan']:
                委托编号 = str(df.loc[idx, "Unnamed: 12"])
        else:
            收货人 = str(df.loc[idx, "Unnamed: 11"])
            if str(df.loc[idx, "Unnamed: 11"]) not in ['', 'nan']:
                委托编号 = str(df.loc[idx, "Unnamed: 11"])

            备注1 = 备注1.replace('：', ':')
            if 备注1.__contains__("备注1:"):
                备注1=备注1.replace(' ','')
                aa=备注1.split("备注1:")
                if(len(aa) == 1):
                    ps1 = aa[0]
                else:
                    ps1 = aa[1]
                while num<itemList.__len__():
                    itemList[num]["备注1"]=ps1
                    num+=1
        
        if 备注1 == "nan": 备注1 = ""
        备注2 = 备注2.replace('：', ':')
        has_correct_备注2 = False
        if 备注2.__contains__("备注2:"):
            has_correct_备注2 = True
            备注2=备注2.replace(' ','')
            aa=备注2.split("备注2:")
            if(len(aa) == 1):
                ps2 = aa[0]
            else:
                ps2 = aa[1]
            while num2<itemList.__len__():
                itemList[num2]["备注2"]=ps2
                num2+=1
        客户组 = 客户组.replace('：', ':')
        if 客户组.__contains__("客户组:"):
            客户组=客户组.replace(' ','')
            aa=客户组.split("客户组:")
            if(len(aa) == 1):
                ps2 = aa[0]
            else:
                ps2 = aa[1]
            while num4<itemList.__len__():
                itemList[num4]["客户组"]=ps2
                itemList[num4]["新办事处"]=新办事处
                num4+=1
        编号 = 编号.replace('：', ':')
        if(编号.__contains__("合同号:")):
            new_contractNo = 编号
            conParts = new_contractNo.split("合同号:")            
            if(len(conParts) == 1):
                new_contractNo = conParts[0]
            else:
                new_contractNo = conParts[1]
            new_contractNo=new_contractNo.replace(" ","")
            new_contractNo=new_contractNo.replace("\n","")
            if contractNo != None and contractNo != new_contractNo:
                raise Exception('录单文件格式有误')
            else:
                contractNo = new_contractNo
            for ite in itemList:
                if ite["合同号"] == "":
                    ite["合同号"] = contractNo
                    #.replace("合同号：","")
        else:
            if(编号 == "nan" and lastId != None):
                编号  = lastId
            else:
                lastId = 编号
            
            if(客户 == "nan" and lastCustomer != None):
                客户  = lastCustomer
            else:
                lastCustomer = 客户

            if(交期 == "nan" and lastDelDate != None):
                交期  = lastDelDate
            else:
                lastDelDate = 交期

            if(收货人 == "nan" and lastReceiver != None):
                收货人  = lastReceiver
            else:
                收货人 = 收货人.replace('\r', '').replace('\n', '')
                lastReceiver = 收货人
            
            if 料号 not in["nan",'是否安装调试验收','', 'N', 'Y'] and not has_correct_备注2:
                if 料号.__contains__("备注"):
                    raise ExcelDataException("1002","备注应与水泵型号对齐")
                beizhu1 = ''
                if has_seperate_beizhu1:
                    beizhu1 = 备注1
                itemList.append({
                    "编号":编号,
                    "客户":客户,
                    "交期":交期 or "",
                    "收货人":收货人,
                    "合同号":"",
                    "料号":料号,
                    "序列号": 序列号列表,
                    "数量":数量,
                    "单价":单价,
                    "工厂":工厂,
                    "配件型号": str(df.loc[idx, "Unnamed: 3"]).strip(),
                    "订单类型":"",
                    "备注1":beizhu1,
                    "备注2":"",
                    "委托编号":委托编号
                })
                if 交期 not in deliDateList:
                    deliDateList.append(交期)

    itemList = parse_result(deliDateList, itemList)

    #根据成品标准周期 判断是否要替换工厂和交期
    #util.update_date(itemList)
    result_items.append(itemList)
    return result_items

def is_zbk(customer_name):
    zkb_names = ['三河同飞制冷股份有限公司', '景津装备股份有限公司', '常德市三一机械有限公司', '三一汽车制造有限公司']
    for name in zkb_names:
        if customer_name.__contains__(name):
            return True
    return False

def JoinFactory(itemList):
    """
    ZUB - 办事处转储采购 ，对应工厂：1072、1073、1079
    ZNB - 跨公司转储     ，对应工厂：3510、3520、1100
    """

    ZUB_FACTORIES = {"1072", "1073", "1079"}
    ZNB_FACTORIES = {"3510", "3520", "1100"}

    log.logger.info(itemList)

    for it in itemList:
        factory = str(it["工厂"]).strip()

        if factory in ZUB_FACTORIES:
            it["订单类型"] = "ZUB"  # 办事处转储采购

        elif factory in ZNB_FACTORIES:
            it["订单类型"] = "ZNB"  # 跨公司转储

        else:
            # 未知工厂
            raise ExcelDataException(
                "1008",
                f"未知工厂类型：{factory}，无法判断订单类型，请核查物料工厂字段"
            )

    return itemList

def GetOrderItems(file):
    result_items = parse(file)
    final_list = []

    for ilist in result_items:
        tem_list = JoinFactory(ilist)
        results = []

        for il in tem_list:
            #  序列号统一转 list
            serial_raw = il.get("序列号", "")
            if isinstance(serial_raw, list):
                serial_list = serial_raw
            elif isinstance(serial_raw, str):
                serial_list = [s.strip() for s in serial_raw.replace("，", ",").split(",") if s.strip()]
            else:
                serial_list = []

            #  单价兜底 
            try:
                il["单价"] = format(float(il["单价"]), ".2f")
            except Exception:
                il["单价"] = "0.00"

            org = "1072"
            新办事处 = il.get("新办事处", "") or ""
            是否安装调试验收 = il.get("是否安装调试验收", "")

            #  查找是否已有同订单 
            exiRows = [
                d for d in results
                if d["编号"] == il["编号"] and d["客户"] == il["客户"]
            ]

            item = {
                "销售组织": org,
                "料号": il["料号"],
                "配件型号": il["配件型号"],
                "序列号": serial_list, 
                "数量": il["数量"],
                "单价": il["单价"],
                "委托编号": il.get("委托编号", ""),
                "工厂": il["工厂"],
                "备注1": il["备注1"],
                "是否安装调试验收": 是否安装调试验收
            }

            if exiRows:
                exiRows[0]["items"].append(item)
            else:
                results.append({
                    "销售组织": org,
                    "新办事处": 新办事处,
                    "编号": il["编号"],
                    "客户": il["客户"],
                    "交期": il["交期"],
                    "收货人": il["收货人"],
                    "合同号": il["合同号"],
                    "订单类型": il["订单类型"],
                    "备注2": il["备注2"],
                    "是否安装调试验收": 是否安装调试验收,
                    "items": [item]
                })

        final_list.extend(results)

    return final_list



def get_network_time():
    response = requests.get('http://www.baidu.com')
    if response.status_code == 200:
        ts = response.headers['date']
        time_arry = time.strptime(ts[5:25], "%d %b %Y %H:%M:%S")
        mytime=time.strftime("%Y.%m.%d",time.localtime(time.mktime(time_arry)+28800))
        return mytime

def getDate(strr):
    mystr=str(strr)
    mystr=mystr.replace('年','.').replace('月','.').replace('日','').replace('/','.')
    return mystr

def parseExport_me2n(file):
    df = pd.read_excel(file, sheet_name=0)
    customer_name=None
    wuliao_num=None
    order_count=None
    #deal_date=None
    deal_address=None
    hetong_num=None
    res=[]
    item_list=[]

    for idx, row in df.iterrows():
        if idx==0:
            deal_address=None
        item_list.append({
            
            "wuliao_num":str(df.loc[idx, row.index[8]]),
            "order_count":str(df.loc[idx, row.index[20]]),
            #TODO 添加产品收货地址
        })

    return {

        "items":item_list
    }

def parseExport(file):#解析sap下载的excel表格EXPORT.XLSX
    df = pd.read_excel(file, sheet_name=0)
    customer_name=None
    wuliao_num=None
    order_count=None
    deal_date=None
    deal_address=None
    hetong_num=None
    res=[]
    item_list=[]

    for idx, row in df.iterrows():
        if idx==0:
            customer_name=str(df.loc[idx, row.index[10]])
            caigou_no=str(df.loc[idx, row.index[14]])
            deal_date=str(df.loc[idx, row.index[28]])
            hetong_num=str(df.loc[idx, row.index[76]])
            deal_address=str(df.loc[idx, row.index[79]])

        item_list.append({
            "wuliao_num":str(df.loc[idx, row.index[18]]),
            "order_count":str(df.loc[idx, row.index[26]]),
        })

    return {
        "deal_date":deal_date,
        "customer_name":customer_name,
        "hetong_num":hetong_num,
        'deal_address':deal_address,
        'caigou_no':caigou_no,
        "items":item_list
    }

def parseSureOrder():
    mytime=get_network_time()
    mytime=mytime.replace('.','-')
    mypath=os.path.join('E:\ORDER\待开单',mytime+'.xlsx')
    res=[]
    if os.path.exists(mypath):
        df = pd.read_excel(mypath, sheet_name='sheet1')
        for idx, row in df.iterrows():
            res.append(str(df.loc[idx, row.index[0]]))

    #aa["entity_list"][0]["meaning"]["value"]

def parseDate(date):#EXPORT.XLSX中Date转换
    mydate=date.split(' ')
    mydate=mydate[0]
    myformat='%Y.%m.%d'
    mydate=mydate.replace('-','.')
    if not mydate.startswith('202'):
        mydate=datetime.datetime.now().strftime('%Y') + '.'+mydate
    dt=datetime.datetime.strptime(mydate, myformat)  
    dt.strftime(myformat)
    return dt.strftime(myformat)  

# def createExcel(mypath,title):
#     '''
#     wb=Workbook()
#     wb.create_sheet('Sheet1')
#     sheet=wb['Sheet1']
#     sheet.append(title)
#     wb.save(mypath)
#     wb.close()
#     '''
#     template_file = os.path.join(os.path.dirname(mypath),"模板.xlsx")
#     shutil.copy(template_file, mypath)

def createExcel(mypath, title):
    os.makedirs(os.path.dirname(mypath), exist_ok=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet.append(title)

    wb.save(mypath)
    wb.close()

def toexcel_id(订单记录_path, id):
    # 统一当天文件名
    date_str = get_network_time().split()[0].replace('.', '-')
    os.makedirs(订单记录_path, exist_ok=True)

    mypath = os.path.join(订单记录_path, f"{date_str}.xlsx")

    # 当天文件不存在则新建
    if not os.path.exists(mypath):
        createExcel(mypath, ['采购订单号', '销售凭证'])

    wb = load_workbook(mypath)
    sheet = wb['Sheet1']
    sheet.append([id, ''])
    wb.save(mypath)
    wb.close()

def _normalize_record_folder(p: str) -> str:
    """
    兼容传入文件路径或文件夹路径：
    - 如果是 xxx.xlsx / xxx.xls：取其 dirname 作为 folder
    - 否则当作 folder
    """
    p = os.path.abspath(p)
    lower = p.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return os.path.dirname(p)
    return p

# ==================== 快照（snapshot）存取：按采购订单号 PO 一单一份 ====================

def _snapshot_dir(订单记录_path: str) -> str:
    base = _normalize_record_folder(订单记录_path)
    p = os.path.join(base, "snapshots")
    os.makedirs(p, exist_ok=True)
    return p


def save_snapshot(订单记录_path: str, 采购订单号: str, order: dict, source_file: str = "") -> str:
    """
    保存录单快照（用于后续 MIGO 等下游流程解耦）
    - 文件名：snapshots/{采购订单号}.json
    - 内容：order（含 items），附加少量元信息
    """
    snap = dict(order) if isinstance(order, dict) else {"order": order}
    snap["_meta"] = {
        "采购订单号": str(采购订单号),
        "source_file": source_file,
        "created_at": get_network_time(),
    }

    path = os.path.join(_snapshot_dir(订单记录_path), f"{采购订单号}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    return path


def load_snapshot(订单记录_path: str, 采购订单号: str) -> dict:
    """
    读取快照：返回 dict；不存在返回 {}
    """
    path = os.path.join(_snapshot_dir(订单记录_path), f"{采购订单号}.json")
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def update_凭证_into_file(订单记录_path, 采购订单号, 销售凭证):
    date_str = get_network_time().split()[0].replace('.', '-')
    mypath = os.path.join(订单记录_path, f"{date_str}.xlsx")

    if not os.path.exists(mypath):
        return  # 没有当天台账直接跳过
    df = pd.read_excel(mypath, dtype=str)
    df.fillna("", inplace=True)

    for idx, row in df.iterrows():
        if row['采购订单号'] == 采购订单号:
            df.loc[idx, '销售凭证'] = 销售凭证

    df.to_excel(mypath, index=False)



def load_订单记录文件(订单记录_path):
    date_str = get_network_time().split()[0].replace('.', '-')
    mypath = os.path.join(订单记录_path, f"{date_str}.xlsx")

    if not os.path.exists(mypath):
        return []
    df = pd.read_excel(mypath, dtype=str, sheet_name='Sheet1')
    df.fillna("", inplace=True)
    return [r for _, r in df.iterrows()]


def getEarlyFile(rootFolder):
    pendingFolder = os.path.join(rootFolder, "待处理")
    files = os.listdir(pendingFolder)
    times=[]
    for f in files:
        mypath=os.path.join(pendingFolder,f)
        times.append(os.path.getmtime(mypath))
    
    myindex=times.index(min(times))
    return files[myindex]

def getEarlyFiles(rootFolder):
    pendingFolder = os.path.join(rootFolder, "待处理")
    files = os.listdir(pendingFolder)
    files.sort(key=lambda fn: os.path.getmtime(pendingFolder + "\\" + fn))
    #files.reverse()
    return files

if __name__ == "__main__":
    #file = r"tmp\华北济南四部2001974614.xlsx"
    file = r"D:\机器人3\测试_20260104水泵退货单4(1).xlsx"
    result = GetOrderItems(file)
    output_path = "发货单数据提取.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"数据提取完成，已保存为：{output_path}")
