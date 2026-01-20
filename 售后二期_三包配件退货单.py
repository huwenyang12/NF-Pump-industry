from openpyxl import load_workbook
import os, json


def _cell_str(v):
    """把单元格值统一转成字符串（None -> ''，数字保持原样输出）"""
    if v is None:
        return ""
    # 避免出现 2.0 / 998.1615000003 这种
    if isinstance(v, float):
        s = f"{v:.10f}".rstrip("0").rstrip(".")
        return s
    return str(v).strip()


def parse_purchase_excel(file_path: str):
    """
    通用解析：
    - 支持只改列字母就加字段
    - 自动分单：供应商编号 + 供应商
    - 输出结构：[{供应商编号, 供应商, items:[...]}]
    """

    HEAD_COLS = {
        "客户参考": "A",
        "售达方编码": "B",
        "CRM号": "C",
    }

    ITEM_COLS = {
        "物料编码": "D",
        "物料名称": "E",
        "数量": "F",
        "单价": "G",
        "金额": "H",
        "销售办事处": "I",
        "类别": "J",
        "工厂": "K",
        "库位": "L",
        "拒绝原因": "M",
    }

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    start_row = 4

    # 1. 先在 D 列定位 “备注2”，并取订单头通用字段 
    stop_row = ws.max_row  # 设置读取的最大行，默认兜底
    is_install = ""
    remark1 = ""
    remark2 = ""

    for r in range(start_row, ws.max_row + 1):
        d_val = _cell_str(ws[f"E{r}"].value)
        # 找到D 列的 “备注2”
        if "备注2" in d_val:
            stop_row = r
            # 用户状态签收
            is_install = _cell_str(ws.cell(row=r, column=4).value)  # r行3列
            # 备注1
            remark1 = _cell_str(ws.cell(row=r - 1, column=9).value)  # r-1行5列
            # 备注2
            remark2 = _cell_str(ws.cell(row=r, column=9).value)  # r行5列
            break

    common_head = {
        "用户签收状态": is_install,
        "备注1": remark1,
        "备注2": remark2,
    }

    results_map = {}  # key -> {供应商编号, 供应商, 是否安装调试, 备注1, 备注2, items:[]}

    # 2. 正式读取数据行：读到 备注2 即 stop_row-1 截至
    for r in range(start_row, stop_row):
        supplier_code = _cell_str(ws[f"{HEAD_COLS['客户参考']}{r}"].value)
        supplier_name = _cell_str(ws[f"{HEAD_COLS['售达方编码']}{r}"].value)

        if not supplier_code and not supplier_name:
            continue
        # 组成分单 Key
        key = f"{supplier_code}_{supplier_name}"

        # 初始化分单
        if key not in results_map:
            results_map[key] = {
                "订单类型": "Z009",
                "销售组织": "1072",
                "分销渠道": "10",
                "产品组": "10",
                "销售办事处": "1000",
                "销售组": "270",
                "客户参考": supplier_code,
                "售达方编码": supplier_name,
                **common_head,   # 个订单头都带上
                "items": []
            }

        item = {}
        for field, col_letter in ITEM_COLS.items():
            item[field] = _cell_str(ws[f"{col_letter}{r}"].value)
        if not item.get("物料编码"):
            continue
        results_map[key]["items"].append(item)

    return list(results_map.values())


if __name__ == "__main__":
    file_path = r"D:\青臣云起\项目\南方流体模板解析\模板\20260115三包配件与电机入库--模板.xlsx"
    data = parse_purchase_excel(file_path)
    out_path = os.path.splitext(file_path)[0] + "_解析结果.json"

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"解析完成，共 {len(data)} 个分单")
    print(f"解析结果写入：{out_path}")
