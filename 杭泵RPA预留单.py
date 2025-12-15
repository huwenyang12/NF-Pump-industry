import json
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

def excel_date_to_str(v):
    if not v:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y.%m.%d")
    s = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}.{s[4:6]}.{s[6:]}"
    return s.replace("-", ".").replace("/", ".")

def is_block_header(ws, row):
    """判断是否为订单区块表头行"""
    return (str(ws[f"A{row}"].value).strip() == "订单类型" and str(ws[f"C{row}"].value).strip() == "物料编号")

def get_merge(ws, cell):
    """读取合并单元格"""
    if ws[cell].value is not None:
        return str(ws[cell].value).strip()
    for m in ws.merged_cells.ranges:
        if cell in m:
            v = ws.cell(m.min_row, m.min_col).value
            return str(v).strip() if v else ""
    return ""


def parse_order_excel(file_path):
    # ---- 格式校验：禁止 xls ----
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        raise ValueError("不支持 .xls 格式，请另存为 .xlsx 后再上传")
    if ext != ".xlsx":
        raise ValueError(f"不支持的文件格式：{ext}，仅支持 .xlsx")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    orders = []
    row = 1

    while row <= ws.max_row:
        # 找到一个区块表头
        if not is_block_header(ws, row):
            row += 1
            continue

        # 当前区块的订单头在表头下一行
        header_row = row + 1

        order = {
            "订单类型": get_merge(ws, f"A{header_row}"),
            "单位名称": get_merge(ws, f"B{header_row}"),
            "采购组织": "3900",
            "采购组": "J04",
            "公司代码": "3900",
            "供货工厂": "3900",
            "items": []
        }

        # 明细从 header_row 开始往下
        r = header_row
        while r <= ws.max_row:
            model = str(ws[f"D{r}"].value or "")

            # 遇到合计当前订单结束
            if "合计" in model:
                break

            material = ws[f"C{r}"].value
            if not material:
                r += 1
                continue

            item = {
                "物料编码": str(material),
                "水泵型号": model,
                "数量": str(ws[f"E{r}"].value or "0"),
                "单价": str(ws[f"F{r}"].value or "0"),
                "金额": str(ws[f"G{r}"].value or "0"),
                "交期": excel_date_to_str(ws[f"H{r}"].value),
                "工厂": str(ws[f"I{r}"].value or "0"),
                "库存地点": str(ws[f"J{r}"].value or ""),
                "营销仓库位号": str(ws[f"K{r}"].value or "")
            }

            order["items"].append(item)
            r += 1

        orders.append(order)

        # row 跳到合计行之后，继续找下一个区块
        row = r + 1

    return orders



if __name__ == "__main__":
    file_path = r"D:\青臣云起\项目\南方流体模板解析\文件\流体东莞办发货通知单-2025-12-15.xlsx"
    result = parse_order_excel(file_path)

    out_path = os.path.join(os.path.dirname(file_path), "json数据解析.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

    print(f"解析完成：{out_path}")
