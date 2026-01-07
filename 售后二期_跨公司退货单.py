import json
import re
import os
from openpyxl import load_workbook


def get_merged_value(ws, cell_ref):
    """读取合并单元格值（取合并区域左上角）"""
    cell = ws[cell_ref]
    if cell.value is not None:
        return str(cell.value).replace("\n", "").strip()
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            val = ws.cell(merged.min_row, merged.min_col).value
            return str(val).replace("\n", "").strip() if val is not None else ""
    return ""


def find_all_block_starts(ws):
    """扫描所有订单块表头行：C列包含 物料号/物料编号"""
    starts = []
    for row in range(1, ws.max_row + 1):
        val = ws[f"C{row}"].value
        if isinstance(val, str) and re.search(r"物料\s*号|物料\s*编\s*号", val):
            starts.append(row)
    return starts


def parse_order_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    all_orders = []
    block_starts = find_all_block_starts(ws)

    for idx, start_row in enumerate(block_starts):
        # 关键：限定本订单块的解析范围，防止串到下一单
        end_row = (block_starts[idx + 1] - 1) if idx < len(block_starts) - 1 else ws.max_row

        # 供应商信息一般在表头下一行（合并单元格）
        info_row = start_row + 1
        bianhao = get_merged_value(ws, f"A{info_row}")
        name = get_merged_value(ws, f"B{info_row}")

        order_data = {
            "订单类型": "跨公司转储",
            "采购组织": "1002",
            "采购组": "W01",
            "公司代码": "1000",
            "供应商编号": bianhao,
            "供应商": name,
            "是否安装调试验收": "",
            "备注1": "",
            "备注2": "",
            "items": []
        }

        row = start_row + 1
        while row <= end_row:
            c_text = get_merged_value(ws, f"C{row}")

            # 结束标记：是否安装调试验收
            if "是否安装调试验收" in c_text:
                order_data["是否安装调试验收"] = get_merged_value(ws, f"C{row + 1}")

                # 备注1 / 备注2：
                for r in range(row, min(row + 6, end_row + 1)):
                    d_val = get_merged_value(ws, f"D{r}")
                    if "备注1" in d_val:
                        order_data["备注1"] = get_merged_value(ws, f"E{r}")
                    if "备注2" in d_val:
                        order_data["备注2"] = get_merged_value(ws, f"E{r}")
                break

            # 跳过空行
            if not c_text:
                row += 1
                continue
            # 避免把表头“物料号”当物料行
            if re.search(r"物料\s*号|物料\s*编\s*号", c_text):
                break
            # 只认“纯数字”的物料号，过滤“合计/客户组”等杂行
            if not re.fullmatch(r"\d{6,30}", c_text):
                row += 1
                continue

            item = {
                "物料号": get_merged_value(ws, f"C{row}"),
                "配件型号": get_merged_value(ws, f"D{row}"),
                "工厂": get_merged_value(ws, f"E{row}"),
                "库存地点": get_merged_value(ws, f"F{row}"),
                "数量": get_merged_value(ws, f"G{row}"),
                "净价": get_merged_value(ws, f"H{row}"),
                "金额": get_merged_value(ws, f"I{row}"),
                "税码": get_merged_value(ws, f"J{row}"),
                "物料文本": get_merged_value(ws, f"K{row}"),
                "物料采购订单文本": get_merged_value(ws, f"L{row}"),
            }
            order_data["items"].append(item)
            row += 1

        # 没有明细就不加（防止误扫到空块）
        if order_data["items"]:
            all_orders.append(order_data)

    return all_orders


if __name__ == "__main__":
    file_path = r"D:\青臣云起\项目\南方流体模板解析\文件\新_跨公司退货单模板.xlsx"
    data = parse_order_excel(file_path)

    json_path = os.path.join(os.path.dirname(file_path), "json数据解析.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    print("解析完成，订单数：", len(data))
